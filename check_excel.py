"""
merge_trades_to_gblast.py
--------------------------
Merges all paper_trades and hybrid_trades CSVs from LIVE-V1 and LIVE-V2
folders into the GBlast_Merged_Report format.

Each trade = 2 consecutive rows:
  Row 1 (BUY)  → entry_time, entry_price, reason=BUY_CALL/BUY_PUT, return=0
  Row 2 (SELL) → exit_time,  exit_price,  reason=TARGET_HIT/STOP_LOSS/EOD, return=total_pnl

Usage:
    python merge_trades_to_gblast.py

Adjust BASE_DIR and VERSION_FOLDERS below to match your paths.
"""

import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIGURATION — update paths as needed
# ─────────────────────────────────────────────
BASE_DIR = r"D:\QUANT_DASHBAORD\trades"

VERSION_FOLDERS = {
    "V1": os.path.join(BASE_DIR, "LIVE - V1"),
    "V2": os.path.join(BASE_DIR, "LIVE - V2"),
}

OUTPUT_FILE = os.path.join(BASE_DIR, "GBlast_Merged_Report_All_Versions.xlsx")

# ─────────────────────────────────────────────
# MAPPINGS
# ─────────────────────────────────────────────
EXIT_REASON_MAP = {
    "TARGET":           "TARGET_HIT",
    "TARGET_HIT":       "TARGET_HIT",
    "STOP_LOSS":        "STOP_LOSS",
    "SL":               "STOP_LOSS",
    "EOD":              "EOD",
    "SIGNAL_REVERSED":  "SIGNAL_REVERSED",
}


def entry_reason(signal_type: str) -> str:
    s = str(signal_type).upper()
    if s in ("BULLISH", "BUY", "BUY_CALL"):
        return "BUY_CALL"
    return "BUY_PUT"


def build_trade_rows(symbol, company, isin, signal_type,
                     entry_time, entry_price,
                     exit_time, exit_price,
                     qty, status, exit_reason_raw, total_pnl, version):
    """
    Returns [BUY_row] for open trades, or [BUY_row, SELL_row] for closed trades.
    signal_b_s is always 'BUY' on entry, 'SELL' on exit — regardless of direction.
    Direction (BULLISH/BEARISH) is captured in the 'reason' column (BUY_CALL / BUY_PUT).
    """
    mapped_exit = EXIT_REASON_MAP.get(str(exit_reason_raw).upper(), str(exit_reason_raw))

    buy_row = {
        "version":         version,
        "name":            symbol,
        "company":         company,
        "trade_date":      entry_time,
        "signal_b_s":      "BUY",
        "triggered_price": entry_price,
        "executed_price":  entry_price,
        "ltp":             entry_price,
        "quantity":        qty,
        "status":          status,
        "isin":            isin,
        "reason":          entry_reason(signal_type),
        "return":          0,
    }

    rows = [buy_row]

    if str(status).upper() == "CLOSED":
        sell_row = {
            "version":         version,
            "name":            symbol,
            "company":         company,
            "trade_date":      exit_time,
            "signal_b_s":      "SELL",
            "triggered_price": exit_price,
            "executed_price":  exit_price,
            "ltp":             exit_price,
            "quantity":        qty,
            "status":          status,
            "isin":            isin,
            "reason":          mapped_exit,
            "return":          total_pnl,
        }
        rows.append(sell_row)

    return rows


def load_paper_trades(csv_path: str, version: str) -> pd.DataFrame:
    df = pd.read_csv(csv_path)
    if df.empty:
        return pd.DataFrame()
    rows = []
    for _, r in df.iterrows():
        rows.extend(build_trade_rows(
            symbol          = str(r.get("tradingsymbol", "")),
            company         = "NIFTY OPTIONS",
            isin            = "",
            signal_type     = str(r.get("signal_type", "")),
            entry_time      = r.get("entry_time", ""),
            entry_price     = r.get("entry_price", 0),
            exit_time       = r.get("exit_time", ""),
            exit_price      = r.get("exit_price", 0),
            qty             = r.get("quantity", 0),
            status          = str(r.get("status", "")),
            exit_reason_raw = r.get("exit_reason", ""),
            total_pnl       = r.get("total_pnl", 0),
            version         = version,
        ))
    return pd.DataFrame(rows)


def load_hybrid_trades(csv_path: str, version: str) -> pd.DataFrame:
    df = pd.read_csv(csv_path)
    if df.empty:
        return pd.DataFrame()
    rows = []
    for _, r in df.iterrows():
        rows.extend(build_trade_rows(
            symbol          = str(r.get("kite_symbol", "")),
            company         = "NIFTY OPTIONS",
            isin            = str(r.get("upstox_instrument_key", "")),
            signal_type     = str(r.get("signal_type", "")),
            entry_time      = r.get("entry_time", ""),
            entry_price     = r.get("entry_price", 0),
            exit_time       = r.get("exit_time", ""),
            exit_price      = r.get("exit_price", 0),
            qty             = r.get("quantity", 0),
            status          = str(r.get("status", "")),
            exit_reason_raw = r.get("exit_reason", ""),
            total_pnl       = r.get("total_pnl", 0),
            version         = version,
        ))
    return pd.DataFrame(rows)


def collect_all_trades() -> pd.DataFrame:
    all_frames = []

    for version, folder in VERSION_FOLDERS.items():
        if not os.path.isdir(folder):
            print(f"[WARN] Folder not found: {folder}")
            continue

        paper_files  = sorted(glob.glob(os.path.join(folder, "paper_trades_*.csv")))
        hybrid_files = sorted(glob.glob(os.path.join(folder, "hybrid_trades_paper_*.csv")))
        print(f"[{version}] Found {len(paper_files)} paper_trades, {len(hybrid_files)} hybrid_trades")

        # Version labels: V1 hybrid → V1.1, V1 paper → V1.2, V2 hybrid → V2.1, V2 paper → V2.2
        hybrid_version = version.replace("V", "V") + ".1"   # e.g. V1.1, V2.1
        paper_version  = version.replace("V", "V") + ".2"   # e.g. V1.2, V2.2

        for f in hybrid_files:
            df = load_hybrid_trades(f, hybrid_version)
            if not df.empty:
                all_frames.append(df)

        for f in paper_files:
            df = load_paper_trades(f, paper_version)
            if not df.empty:
                all_frames.append(df)

    if not all_frames:
        print("[ERROR] No data found. Check VERSION_FOLDERS paths.")
        return pd.DataFrame()

    combined = pd.concat(all_frames, ignore_index=True)

    # Assign a trade_group_id so BUY/SELL pairs always stay together after sort.
    # BUY rows have no SELL row before them, so stable sort on entry date keeps pairs intact.
    combined.sort_values(["version", "trade_date"], kind="stable", inplace=True)
    combined.reset_index(drop=True, inplace=True)
    return combined


def write_excel(df: pd.DataFrame, output_path: str):
    COLUMNS = ["version", "name", "company", "trade_date", "signal_b_s",
               "triggered_price", "executed_price", "ltp", "quantity",
               "status", "isin", "reason", "return"]

    wb = Workbook()
    ws = wb.active
    ws.title = "GBlast All Versions"

    # ── Header ──
    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="AAAAAA")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper().replace("_", " "))
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border
    ws.row_dimensions[1].height = 22

    # BUY = lighter shade, SELL = slightly darker shade per version
    # V1.1/V1.2 share blue family; V2.1/V2.2 share green family
    buy_colors  = {
        "V1.1": "D6E8FF", "V1.2": "C2DBFF",
        "V2.1": "D4EDDA", "V2.2": "BEE5C8",
        "V3.1": "FFF3CD", "V3.2": "FFE9A0",
        "V4.1": "FAE0E4", "V4.2": "F5C6CE",
    }
    sell_colors = {
        "V1.1": "B8D4F5", "V1.2": "A0C4F0",
        "V2.1": "B2DFC0", "V2.2": "9AD5AC",
        "V3.1": "FFE599", "V3.2": "FFD966",
        "V4.1": "F5C6CE", "V4.2": "EFAAB5",
    }
    buy_font    = Font(name="Arial", size=9)
    sell_font   = Font(name="Arial", size=9, italic=True)

    for row_idx, (_, row) in enumerate(df[COLUMNS].iterrows(), start=2):
        version  = str(row["version"])
        is_sell  = str(row["signal_b_s"]).upper() == "SELL"
        bg       = (sell_colors if is_sell else buy_colors).get(version, "F5F5F5")
        fill     = PatternFill("solid", start_color=bg, end_color=bg)
        rfont    = sell_font if is_sell else buy_font

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            val  = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = rfont; cell.fill = fill; cell.border = border
            if col_name in ("triggered_price", "executed_price", "ltp", "return"):
                cell.number_format = "#,##0.00"
            elif col_name == "quantity":
                cell.number_format = "#,##0"

    col_widths = [8, 28, 16, 24, 10, 16, 16, 12, 10, 10, 20, 18, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    print(f"\n[OK] Saved → {output_path}")
    print(f"     {len(df)} rows  |  ~{len(df[df['signal_b_s']=='BUY'])} trades")


def main():
    print("Collecting trades...\n")
    df = collect_all_trades()
    if df.empty:
        return
    print(f"\nRow breakdown:")
    print(df.groupby(["version", "signal_b_s"]).size().to_string())
    write_excel(df, OUTPUT_FILE)


if __name__ == "__main__":
    main()