import os
import glob
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TRADES_DIR = r"d:\QUANT_DASHBAORD\trades"
OUTPUT_DIR = r"d:\QUANT_DASHBAORD\BLAZE TRADES"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─────────────────────────── helpers ────────────────────────────

def style_sheet(ws, df):
    """Apply professional formatting to a worksheet."""
    # Header fill / font
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(bold=True, color="E0E0FF", size=11)
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Alternate row shading + P&L colouring
    green_fill  = PatternFill("solid", fgColor="D6F5D6")
    red_fill    = PatternFill("solid", fgColor="FFD6D6")
    alt_fill    = PatternFill("solid", fgColor="F0F0F8")
    white_fill  = PatternFill("solid", fgColor="FFFFFF")

    pnl_cols = [i for i, c in enumerate(df.columns, 1)
                if 'pnl' in c.lower() or 'profit' in c.lower()]

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
        is_alt = row_idx % 2 == 0
        pnl_value = None

        # Try to detect P&L from total_pnl column
        for cell in row:
            if ws.cell(row=1, column=cell.column).value in ('total_pnl', 'pnl', 'P&L'):
                try:
                    pnl_value = float(str(cell.value).replace(',', ''))
                except:
                    pass
                break

        for cell in row:
            col_idx = cell.column
            if col_idx in pnl_cols and pnl_value is not None:
                cell.fill = green_fill if pnl_value >= 0 else red_fill
            else:
                cell.fill = alt_fill if is_alt else white_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    # Auto-fit columns
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0 for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

    # Freeze header
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions


def add_summary_sheet(writer, df, version_name):
    """Add a Summary sheet with key stats."""
    if df.empty:
        return
    pnl_col = next((c for c in df.columns if c.lower() == 'total_pnl'), None)
    pct_col  = next((c for c in df.columns if c.lower() == 'pnl_pct'),   None)
    exit_col = next((c for c in df.columns if c.lower() == 'exit_reason'), None)

    rows = []
    rows.append(["Metric", "Value"])
    rows.append(["Strategy", version_name])
    rows.append(["Total Trades", len(df)])

    if pnl_col:
        pnl = pd.to_numeric(df[pnl_col], errors='coerce')
        rows.append(["Total P&L (Rs.)", round(pnl.sum(), 2)])
        rows.append(["Winning Trades", int((pnl > 0).sum())])
        rows.append(["Losing Trades",  int((pnl < 0).sum())])
        rows.append(["Win Rate (%)",   round((pnl > 0).sum() / len(pnl) * 100, 2)])
        rows.append(["Best Trade (Rs.)",  round(pnl.max(), 2)])
        rows.append(["Worst Trade (Rs.)", round(pnl.min(), 2)])
        rows.append(["Avg P&L / Trade (Rs.)", round(pnl.mean(), 2)])

    if pct_col:
        pct = pd.to_numeric(df[pct_col].astype(str).str.replace('%','').str.replace('+',''), errors='coerce')
        rows.append(["Avg Return / Trade (%)", round(pct.mean(), 2)])

    if exit_col:
        ec = df[exit_col].value_counts()
        for reason, count in ec.items():
            rows.append([f"Exit: {reason}", count])

    summary_df = pd.DataFrame(rows[1:], columns=rows[0])
    summary_df.to_excel(writer, sheet_name="Summary", index=False)

    ws = writer.sheets["Summary"]
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(bold=True, color="E0E0FF", size=11)
    key_font = Font(bold=True, size=10, color="1A1A2E")
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in range(1, 3):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx in range(2, ws.max_row + 1):
        key_cell = ws.cell(row=row_idx, column=1)
        val_cell = ws.cell(row=row_idx, column=2)
        key_cell.font = key_font
        key_cell.alignment = Alignment(horizontal="left")
        val_cell.alignment = Alignment(horizontal="right")

        for cell in (key_cell, val_cell):
            cell.border = border

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20


# ──────────────────────── file categorisation ────────────────────

def get_files(pattern_include, pattern_exclude=None):
    files = []
    # Use recursive globbing
    full_pattern = os.path.join(TRADES_DIR, "**", pattern_include)
    for f in glob.glob(full_pattern, recursive=True):
        if pattern_exclude and pattern_exclude in os.path.basename(f):
            continue
        files.append(f)
    return sorted(files)

# ──────────────────────── file categorisation ────────────────────

def get_files_recursively(folder_name, pattern="*.csv"):
    folder_path = os.path.join(TRADES_DIR, folder_name)
    if not os.path.exists(folder_path):
        return []
    return sorted(glob.glob(os.path.join(folder_path, "**", pattern), recursive=True))

# Map versions to their new dedicated folders
v1_files   = get_files_recursively("Blaze v1 - v1")
v2_files   = get_files_recursively("Blaze v2 -v2")
v3_files   = get_files_recursively("Blaze v3 - v3")
v4_files   = get_files_recursively("Blaze v4 - v4")
v42_files  = get_files_recursively("Blaze v4.2 - v4.2")
v5_files   = get_files_recursively("Blaze 5 - v5")
b20_files  = get_files_recursively("B - 20 - Nifty BEES")

# ─────────────────────────── build Excels ────────────────────────

versions = [
    ("Blaze V1",         v1_files),
    ("Blaze V2",         v2_files),
    ("Blaze V3",         v3_files),
    ("Blaze V4",         v4_files),
    ("Blaze V4.2",       v42_files),
    ("Blaze V5",         v5_files),
    ("B-20 (NiftyBEES)", b20_files),
]

for version_name, files in versions:
    if not files:
        print(f"[SKIP] {version_name} — no files found.")
        continue

    dfs = []
    for f in files:
        try:
            df_tmp = pd.read_csv(f)
            dfs.append(df_tmp)
        except Exception as e:
            print(f"  Warning: could not read {f}: {e}")

    if not dfs:
        print(f"[SKIP] {version_name} — no readable CSVs.")
        continue

    combined = pd.concat(dfs, ignore_index=True)

    # Remove instrument_key if requested
    if 'instrument_key' in combined.columns:
        combined.drop(columns=['instrument_key'], inplace=True)

    # Sort by entry_time if available
    if 'entry_time' in combined.columns:
        combined['entry_time'] = pd.to_datetime(combined['entry_time'], errors='coerce')
        combined.sort_values('entry_time', inplace=True)
        combined['entry_time'] = combined['entry_time'].dt.strftime('%Y-%m-%d %H:%M:%S')
    if 'exit_time' in combined.columns:
        combined['exit_time'] = pd.to_datetime(combined['exit_time'], errors='coerce')
        combined['exit_time'] = combined['exit_time'].dt.strftime('%Y-%m-%d %H:%M:%S')

    safe_name = version_name.replace("/", "-").replace("\\", "-")
    out_path = os.path.join(OUTPUT_DIR, f"{safe_name}.xlsx")

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        add_summary_sheet(writer, combined, version_name)
        combined.to_excel(writer, sheet_name="All Trades", index=False)
        ws = writer.sheets["All Trades"]
        style_sheet(ws, combined)

    print(f"[OK]   {version_name}: {len(combined)} trades → {out_path}")

print("\nDone! All Excel files created in:", OUTPUT_DIR)
